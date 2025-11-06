import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

print("="*80)
print("FINAL 'nan' FIX - DIRECT CELL MANIPULATION")
print("="*80)

# Fix the FHM file directly at cell level
file_path = 'Leads - FHM 2025_Updated_20251016_094206.xlsx'
wb = openpyxl.load_workbook(file_path)

sheet_name = 'Consolidated Leads'
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    print(f"\nProcessing sheet: {sheet_name}")
    print(f"Total rows: {ws.max_row}")
    
    nan_count = 0
    fixed_count = 0
    
    # Iterate through all cells
    for row in ws.iter_rows(min_row=2):  # Skip header
        for cell in row:
            if cell.value is not None:
                # Convert to string and check
                cell_str = str(cell.value).strip()
                if cell_str.lower() == 'nan' or cell_str == '':
                    cell.value = ''  # Set to empty string
                    fixed_count += 1
                else:
                    # Keep the value as-is
                    pass
    
    print(f"✓ Fixed {fixed_count} cells with 'nan' or empty values")
    
    # Get column indices for Visitor Name and Mobile No.
    visitor_name_col = None
    mobile_col = None
    
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Visitor Name':
            visitor_name_col = col_idx
        elif cell.value == 'Mobile No.':
            mobile_col = col_idx
    
    print(f"\nVisitor Name column: {visitor_name_col}")
    print(f"Mobile No. column: {mobile_col}")
    
    # Extra pass specifically for these columns
    if visitor_name_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=visitor_name_col)
            if cell.value:
                if str(cell.value).strip().lower() == 'nan':
                    cell.value = ''
    
    if mobile_col:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=mobile_col)
            if cell.value:
                cell_str = str(cell.value).strip()
                if cell_str.lower() == 'nan':
                    cell.value = ''
                elif cell_str and cell_str[0].isdigit():
                    # Ensure space prefix for numbers
                    if not str(cell.value).startswith(' '):
                        cell.value = ' ' + cell_str

# Save the workbook
output_file = 'Leads - FHM 2025_FINAL_20251016.xlsx'
wb.save(output_file)

print(f"\n{'='*80}")
print(f"SAVED: {output_file}")
print(f"{'='*80}")

# Verify by reading back
print(f"\nVerifying...")
df_verify = pd.read_excel(output_file, sheet_name='Consolidated Leads')

# Check Visitor Name
visitor_nan = 0
for val in df_verify['Visitor Name']:
    if pd.notna(val) and 'nan' in str(val).lower():
        visitor_nan += 1

print(f"\n'nan' in Visitor Name: {visitor_nan}")

# Check Mobile No.
mobile_nan = 0
for val in df_verify['Mobile No.']:
    if pd.notna(val) and 'nan' in str(val).lower():
        mobile_nan += 1

print(f"'nan' in Mobile No.: {mobile_nan}")

# Show samples
print(f"\n{'='*80}")
print("SAMPLE DATA FROM CLEANED FILE:")
print(f"{'='*80}")

sample = df_verify.head(10)
for idx, row in sample.iterrows():
    name = row['Visitor Name'] if pd.notna(row['Visitor Name']) and str(row['Visitor Name']).strip() else '[Empty]'
    mobile = row['Mobile No.'] if pd.notna(row['Mobile No.']) and str(row['Mobile No.']).strip() else '[Empty]'
    company = row['Visitor Company'] if pd.notna(row['Visitor Company']) and str(row['Visitor Company']).strip() else '[Empty]'
    
    print(f"\nRow {idx+1}:")
    print(f"  Name: {name}")
    print(f"  Company: {company}")
    print(f"  Mobile: {mobile}")

if visitor_nan == 0 and mobile_nan == 0:
    print(f"\n{'='*80}")
    print("✓✓✓ SUCCESS! NO 'nan' IN VISITOR NAME OR MOBILE NO.! ✓✓✓")
    print(f"{'='*80}")
else:
    print(f"\n⚠ Still found issues - need further investigation")

print(f"\n{'='*80}")
print("FINAL FILES:")
print(f"{'='*80}")
print(f"\n1. For Hubspot: Hubspot_Import_FINAL_20251016.csv")
print(f"2. For FHM 2025: {output_file}")
print()


